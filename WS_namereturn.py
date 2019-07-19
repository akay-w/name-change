# -*- coding: utf-8 -*-
"""
Created on Tue Mar 12 13:38:51 2019

@author: a-whalen
"""

import tkinter as tk
from tkinter.filedialog import askdirectory, askopenfilename
import os
import openpyxl
root = tk.Tk()
root.withdraw()

filelist = askopenfilename()
wb = openpyxl.load_workbook(filelist)
ws = wb.active

origcol = 1
editcol = 2
maxRow = ws.max_row
file_dict = {}
for i in range(2, maxRow+1):
    origfile = ws.cell(row = i, column = origcol).value
    editfile = ws.cell(row = i, column = editcol).value
    file_dict[editfile] = origfile
    
directory = askdirectory()
filelist = []
for root, dirs, files in os.walk(directory):
    for file in files:
        if file.endswith("sdlxliff"):
            filelist.append(os.path.join(os.path.normpath(root), file))

for filepath in filelist:
    editfile = os.path.basename(filepath)
    origfile = file_dict[editfile]
    newfilepath = filepath.replace(editfile, origfile)
    os.rename(filepath, newfilepath)